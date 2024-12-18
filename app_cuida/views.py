from .models import Paciente, Especialidade, Medico, Consulta, Events
from .models import Biometria, SinaisVitais, CondicoesEspeciais, Alergia, MedicamentoAtivo
from django.contrib.auth import authenticate, login as auth_login
from django.contrib.auth.models import User
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib import messages
from django.db.models import Q
from django.http import HttpResponse, JsonResponse, HttpResponseForbidden
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse
from datetime import datetime, timedelta
import pandas as pd
import openpyxl
import calendar


@login_required(login_url='login')
def add(request):
    if request.method == 'POST':
        nome = request.POST.get('nome')
        idade = request.POST.get('idade')
        numero_celular = request.POST.get('numero_celular')
        numero_prontuario = request.POST.get('numero_prontuario')
        sexo = request.POST.get('sexo')
        cpf = request.POST.get('cpf')
        status = request.POST.get('status')

        if Paciente.objects.filter(Q(cpf=cpf) | Q(numero_prontuario=numero_prontuario)).exists():
            messages.error(request, 'Já existe um paciente com o mesmo CPF ou número de prontuário.')
            return redirect('form')

        paciente = Paciente(
            nome=nome,
            idade=idade,
            numero_celular=numero_celular,
            numero_prontuario=numero_prontuario,
            sexo=sexo,
            cpf=cpf,
            status=status
        )

        try:
            paciente.save()
            messages.success(request, 'Paciente cadastrado com sucesso!')
            return redirect('home')
        except:
            messages.error(request, 'Ocorreu um erro ao cadastrar o paciente, por favor tente de novo.')

    return render(request, 'cadastro/form.html')

@login_required(login_url='login')
def update(request, id_paciente):
    paciente = Paciente.objects.filter(id_paciente=id_paciente).first()
    
    if request.method == 'POST':
        paciente.nome = request.POST.get('nome')
        paciente.idade = request.POST.get('idade')
        paciente.numero_celular = request.POST.get('numero_celular')
        paciente.numero_prontuario = request.POST.get('numero_prontuario')
        paciente.sexo = request.POST.get('sexo')
        paciente.cpf = request.POST.get('cpf')
        paciente.status = request.POST.get('status')

        paciente.save()
        return redirect('listagem_pacientes')

    context = {
        'paciente': paciente
    }
    return render(request, 'cadastro/form.html', context)

@login_required(login_url='login')
def visualizar(request):
    pacientes = Paciente.objects.all()
    context = {
        'pacientes': pacientes
    }
    return render(request, 'cadastro/pacientes.html', context)

@login_required(login_url='login')
def visualizar_edit(request):
    pacientes = Paciente.objects.all()
    context = {
        'pacientes': pacientes
    }
    return render(request, 'cadastro/editar.html', context)

@login_required(login_url='login')
def delete_paciente(request, id_paciente):
    paciente = get_object_or_404(Paciente, id_paciente=id_paciente)
    paciente.delete()
    return redirect('listagem_pacientes')

def cadastro(request):
    if request.method == 'GET':
        return render(request, 'cadastro/login.html')
    else:
        username = request.POST.get('username')
        email = request.POST.get('email')
        senha = request.POST.get('senha')

        user = User.objects.filter(username=username).first()

        if user:
            return HttpResponse('Já existe um usuário com esse nome!')
        
        user = User.objects.create_user(username=username, email=email, password=senha)
        user.save()

        return render(request, 'cadastro/login.html', {'success_message': 'Usuário cadastrado com sucesso!'})
    
def is_admin(user):
    return user.is_superuser
    
def cadastrar_especialidade(request):
    if request.method == 'POST':
        nome = request.POST.get('nome')

        if nome:
            if not Especialidade.objects.filter(nome=nome).exists():
                Especialidade.objects.create(nome=nome)
                return redirect('visualizar_especialidades')
            else:
                messages.error(request, "Essa especialidade já está cadastrada.")
                return redirect('cadastrar_especialidade') 

    return render(request, 'cadastro/cadastrar_especialidade.html')

def visualizar_especialidades(request):
    especialidades = Especialidade.objects.all()
    context = {
        'especialidades': especialidades
    }
    return render(request, 'cadastro/lista_especialidades.html', context)


def excluir_especialidade(request, id):
    if request.method == 'POST':
        especialidade = get_object_or_404(Especialidade, id=id)
        especialidade.delete()
        return redirect('visualizar_especialidades')

def cadastrar_medico(request):
    if request.method == 'POST':
        nome = request.POST.get('nome')
        especialidade_id = request.POST.get('especialidade')
        crm = request.POST.get('crm')
        numero_celular = request.POST.get('numero_celular')

        if nome and especialidade_id and crm:
            especialidade = get_object_or_404(Especialidade, id=especialidade_id)

            if Medico.objects.filter(crm=crm).exists():
                messages.error(request, "Um médico com esse CRM já está cadastrado.")
                return redirect('cadastrar_medico')

            Medico.objects.create(nome=nome, especialidade=especialidade, crm=crm, numero_celular=numero_celular)
            return redirect('visualizar_medicos')

    especialidades = Especialidade.objects.all()
    return render(request, 'cadastro/cadastrar_medico.html', {'especialidades': especialidades})

def visualizar_medicos(request):
    medicos = Medico.objects.all()
    context = {
        'medicos': medicos
    }
    return render(request, 'cadastro/lista_medicos.html', context)

def excluir_medico(request, id):
    medico = get_object_or_404(Medico, id=id)  
    if request.method == 'POST':
        medico.delete() 
        return redirect('visualizar_medicos') 
    return render(request, 'confirmar_exclusao.html', {'medico': medico})

def cadastrar_consulta(request):
    if request.method == 'POST':
        id_paciente = request.POST.get('paciente')
        medico_id = request.POST.get('medico')
        data_consulta_str = request.POST.get('data_consulta')
        horario = request.POST.get('horario')

        print(f"Paciente ID: {id_paciente}, Médico ID: {medico_id}, Data: {data_consulta_str}, Horário: {horario}")

        if not id_paciente or not medico_id or not data_consulta_str or not horario:
            messages.error(request, "Por favor, preencha todos os campos obrigatórios.")
            return render(request, 'cadastro/cadastrar_consulta.html', {
                'pacientes': Paciente.objects.all(),
                'medicos': Medico.objects.all(),
                'horarios': ['08:00', '09:00', '10:00', '14:00', '15:00'],
            })

        paciente = Paciente.objects.get(id_paciente=id_paciente)
        medico = Medico.objects.get(id=medico_id)
        data_consulta = datetime.strptime(data_consulta_str, '%Y-%m-%d').date()

        if data_consulta < datetime.now().date():
            messages.error(request, "A data da consulta não pode ser no passado.")
            return render(request, 'cadastro/cadastrar_consulta.html', {
                'pacientes': Paciente.objects.all(),
                'medicos': Medico.objects.all(),
                'horarios': ['08:00', '09:00', '10:00', '14:00', '15:00'],
            })

        if Consulta.objects.filter(medico=medico, data_consulta=data_consulta, horario=horario).exists():
            messages.error(request, "O médico já está ocupado neste horário.")
            return render(request, 'cadastro/cadastrar_consulta.html', {
                'pacientes': Paciente.objects.all(),
                'medicos': Medico.objects.all(),
                'horarios': ['08:00', '09:00', '10:00', '14:00', '15:00'],
            })

        Consulta.objects.create(paciente=paciente, medico=medico, data_consulta=data_consulta, horario=horario)

        start = f"{data_consulta_str} {horario}"
        end = f"{data_consulta_str} {horario}"
        title = f"{medico.nome} - {paciente.nome}"
        event = Events(name=title, start=start, end=end)
        event.save()

        return redirect('visualizar_consultas')

    pacientes = Paciente.objects.all()
    medicos = Medico.objects.all()
    horarios = ['08:00', '09:00', '10:00', '14:00', '15:00']

    return render(request, 'cadastro/cadastrar_consulta.html', {
        'pacientes': pacientes,
        'medicos': medicos,
        'horarios': horarios,
    })

def visualizar_consultas(request):
    consultas = Consulta.objects.all()
    context = {
        'consultas': consultas
    }
    return render(request, 'cadastro/lista_consultas.html', context)

def excluir_consulta(request, id):
    consulta = get_object_or_404(Consulta, id=id)
    consulta.delete()
    return redirect('visualizar_consultas')

def calendario_view(request):
    # Recupera o mês e ano da URL ou usa o atual
    month = int(request.GET.get('month', datetime.now().month))
    year = int(request.GET.get('year', datetime.now().year))

    # Ajusta o mês e ano caso estejam fora do intervalo permitido
    if month < 1:
        month = 12
        year -= 1
    elif month > 12:
        month = 1
        year += 1

    # Primeiro dia do mês e número de dias no mês
    first_day = datetime(year, month, 1)
    next_month = datetime(year + 1, 1, 1) if month == 12 else datetime(year, month + 1, 1)
    num_days = (next_month - first_day).days
    days = [first_day + timedelta(days=i) for i in range(num_days)]

    # Consulta as consultas para o mês atual
    consultas = Consulta.objects.filter(data_consulta__year=year, data_consulta__month=month)
    consultas_por_dia = {}
    for consulta in consultas:
        dia = consulta.data_consulta
        if dia not in consultas_por_dia:
            consultas_por_dia[dia] = []
        consultas_por_dia[dia].append(consulta)

    # Nomes dos meses
    month_name = first_day.strftime('%B')

    context = {
        'days': days,
        'consultas_por_dia': consultas_por_dia,
        'month': month,
        'year': year,
        'month_name': month_name,
    }
    return render(request, 'cadastro/calendario.html', context)

def login(request):
    if request.method == 'GET':
        return render(request, 'cadastro/login.html')
    else:
        username = request.POST.get('username')
        senha = request.POST.get('senha')

        # Autentica o usuário
        user = authenticate(request, username=username, password=senha)

        if user is not None:
            auth_login(request, user)
            return redirect('home')
        else:
            return HttpResponse('Usuário ou senha inválidos!')

@login_required(login_url='login')
def home(request):
    return render(request, 'cadastro/home.html')

def detalhes_paciente(request, id_paciente):
    paciente = get_object_or_404(Paciente, id_paciente=id_paciente)
    biometrias = Biometria.objects.filter(paciente=paciente).order_by('-data_consulta')[:2]
    sinais_vitais = SinaisVitais.objects.filter(paciente=paciente).order_by('-data_consulta')[:2]
    condicoes_especiais = CondicoesEspeciais.objects.filter(paciente=paciente)
    alergias = Alergia.objects.filter(paciente=paciente)
    medicamentos_ativos = MedicamentoAtivo.objects.filter(paciente=paciente)

    context = {
        'paciente': paciente,
        'biometrias': biometrias,
        'sinais_vitais': sinais_vitais,
        'condicoes_especiais': condicoes_especiais,
        'alergias': alergias,
        'medicamentos_ativos': medicamentos_ativos,
    }
    return render(request, 'cadastro/detalhes_paciente.html', context)


def index(request):  
    all_consultas = Consulta.objects.all()
    context = {
        "consultas": all_consultas,
    }
    return render(request, 'cadastro/index.html', context)

def all_consultas(request):
    consultas = Events.objects.all()
    events = []

    for consulta in consultas:
        events.append({
            'title': consulta.name,
            'id': consulta.id,
            'start': consulta.start.strftime("%m/%d/%Y, %H:%M:%S"),
            'end': consulta.end.strftime("%m/%d/%Y, %H:%M:%S"),
        })
    
    data = {
        'events': events
    }

def gerar_relatorio(request):
    try:
        pacientes = Paciente.objects.all().values(
            'id_paciente', 'nome', 'idade', 'cpf', 'numero_celular', 
            'numero_prontuario', 'sexo', 'status'
        )
        
        df = pd.DataFrame(pacientes)
        
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename=relatorio_pacientes.xlsx'
        
        df.to_excel(response, index=False)
        
        return response
    except:
        return HttpResponse("Erro ao gerar o relatório, tente novamente mais tarde", content_type="text/plain")

def adicionar_biometria(request, id_paciente):
    paciente = get_object_or_404(Paciente, id_paciente=id_paciente)
    
    if request.method == 'POST':
        data_consulta = request.POST.get('data_consulta')
        peso = request.POST.get('peso')
        altura = request.POST.get('altura')
        
        # Verificar se os campos estão preenchidos
        if data_consulta and peso and altura:
            Biometria.objects.create(
                paciente=paciente,
                data_consulta=data_consulta,
                peso=peso,
                altura=altura
            )
            messages.success(request, 'Biometria adicionada com sucesso!')
            return redirect('detalhes_paciente', id_paciente=paciente.id_paciente)
        else:
            messages.error(request, 'Por favor, preencha todos os campos.')
    
    return redirect('cadastro/detalhes_paciente', id_paciente=paciente.id_paciente)

@login_required(login_url='login')
def excluir_biometria(request, id_biometria, id_paciente):
    biometria = get_object_or_404(Biometria, id=id_biometria)
    biometria.delete()
    messages.success(request, 'Biometria excluída com sucesso!')
    return redirect('detalhes_paciente', id_paciente=id_paciente)

@login_required(login_url='login')
def adicionar_sinais_vitais(request, id_paciente):
    paciente = get_object_or_404(Paciente, id_paciente=id_paciente)
    
    if request.method == 'POST':
        data_consulta = request.POST.get('data_consulta')
        temperatura = request.POST.get('temperatura')
        pulso = request.POST.get('pulso')
        pressao = request.POST.get('pressao')
        
        # Verificar se os campos estão preenchidos
        if data_consulta and temperatura and pulso and pressao:
            SinaisVitais.objects.create(
                paciente=paciente,
                data_consulta=data_consulta,
                temperatura=temperatura,
                pulso=pulso,
                pressao=pressao
            )
            messages.success(request, 'Sinais vitais adicionados com sucesso!')
            return redirect('detalhes_paciente', id_paciente=paciente.id_paciente)
        else:
            messages.error(request, 'Por favor, preencha todos os campos.')
    
    return redirect('cadastro/detalhes_paciente', id_paciente=paciente.id_paciente)

@login_required(login_url='login')
def excluir_sinais_vitais(request, id_sinais_vitais, id_paciente):
    sinais_vitais = get_object_or_404(SinaisVitais, id=id_sinais_vitais)
    sinais_vitais.delete()
    messages.success(request, 'Sinais vitais excluídos com sucesso!')
    return redirect('detalhes_paciente', id_paciente=id_paciente)

def adicionar_condicao_especial(request, id_paciente):
    paciente = get_object_or_404(Paciente, id_paciente=id_paciente)
    if request.method == 'POST':
        descricao = request.POST['descricao']
        CondicoesEspeciais.objects.create(paciente=paciente, descricao=descricao)
        return redirect(reverse('detalhes_paciente', args=[id_paciente]))
    return render(request, 'detalhes_paciente.html', {'paciente': paciente})

def adicionar_alergia(request, id_paciente):
    paciente = get_object_or_404(Paciente, id_paciente=id_paciente)
    if request.method == 'POST':
        descricao = request.POST['descricao']
        Alergia.objects.create(paciente=paciente, descricao=descricao)
        return redirect(reverse('detalhes_paciente', args=[id_paciente]))
    return render(request, 'detalhes_paciente.html', {'paciente': paciente})

def adicionar_medicamento_ativo(request, id_paciente):
    paciente = get_object_or_404(Paciente, id_paciente=id_paciente)
    if request.method == 'POST':
        descricao = request.POST['descricao']
        MedicamentoAtivo.objects.create(paciente=paciente, descricao=descricao)
        return redirect(reverse('detalhes_paciente', args=[id_paciente]))
    return render(request, 'detalhes_paciente.html', {'paciente': paciente})

@login_required(login_url='login')
def excluir_condicao_especial(request, id_condicao, id_paciente):
    condicao = get_object_or_404(CondicoesEspeciais, id=id_condicao)
    condicao.delete()
    messages.success(request, 'Condição especial excluída com sucesso!')
    return redirect('detalhes_paciente', id_paciente=id_paciente)

@login_required(login_url='login')
def excluir_alergia(request, id_alergia, id_paciente):
    alergia = get_object_or_404(Alergia, id=id_alergia)
    alergia.delete()
    messages.success(request, 'Alergia excluída com sucesso!')
    return redirect('detalhes_paciente', id_paciente=id_paciente)

@login_required(login_url='login')
def excluir_medicamento_ativo(request, id_medicamento, id_paciente):
    medicamento = get_object_or_404(MedicamentoAtivo, id=id_medicamento)
    medicamento.delete()
    messages.success(request, 'Medicamento ativo excluído com sucesso!')
    return redirect('detalhes_paciente', id_paciente=id_paciente)

def exportar_paciente_excel(request, id_paciente):
    paciente = Paciente.objects.get(id_paciente=id_paciente)
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Detalhes do Paciente"
    
    sheet["A1"] = "Nome"
    sheet["B1"] = paciente.nome
    sheet["A2"] = "Sexo"
    sheet["B2"] = paciente.get_sexo_display()
    sheet["A3"] = "Idade"
    sheet["B3"] = paciente.idade
    sheet["A4"] = "Contato"
    sheet["B4"] = paciente.numero_celular
    
    biometria_sheet = workbook.create_sheet(title="Biometria")
    biometria_sheet.append(["Data", "Peso (kg)", "Altura (cm)", "IMC"])
    for bio in Biometria.objects.filter(paciente=paciente):
        biometria_sheet.append([bio.data_consulta, bio.peso, bio.altura, bio.imc])
    
    sinais_vitais_sheet = workbook.create_sheet(title="Sinais Vitais")
    sinais_vitais_sheet.append(["Data", "Temperatura (°C)", "Pulso (BPM)", "Pressão (mmHg)"])
    for sinais in SinaisVitais.objects.filter(paciente=paciente):
        sinais_vitais_sheet.append([sinais.data_consulta, sinais.temperatura, sinais.pulso, sinais.pressao])
    
    condicoes_sheet = workbook.create_sheet(title="Condições Especiais")
    condicoes_sheet.append(["Descrição"])
    for condicao in CondicoesEspeciais.objects.filter(paciente=paciente):
        condicoes_sheet.append([condicao.descricao])
    
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{paciente.nome}_detalhes.xlsx"'
    workbook.save(response)
    return response